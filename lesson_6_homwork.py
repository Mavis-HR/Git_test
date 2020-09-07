"""
=====================================
Author:yaomeiwen
Time:2020/9/4

=====================================

"""
"""
接口测试自动化的步骤：
1.接口测试用例---
2.Python代码读取接口测试用例
3.requests库发送接口请求
4.执行结果与预期结果比对，写入通过或不通过
5.结果回写到excel里---openpyxl






代码自动读取测试数据 + 自动化回写数据 === 测试用例一般excel里居多 == 操作excel
第三方库：openpyxl --- 读取，回写
1.安装 pip install openpyxl
2.导入
函数定义：
1.实现功能2.参数--变化的值 3.返回值--别人需要从你这里得到的数据




excel表格的常用操作

"""
import requests
import openpyxl # 导入第三方库 openpyxl

def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)  #加载工作对象
    sheet = wb[sheetname] # 获取到表单
    case_list = []
    max_row = sheet.max_row
    for i in range(2,max_row+1):
        case = dict(
        case_id =sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,  #行，列 找到对应的单元格
        data =sheet.cell(row=i,column=6).value,  # 参数
        expected = sheet.cell(row=i,column=7).value
        )   #一个字典是一个测试用例
        case_list.append(case)
    return case_list     #返回值

cases = read_data("test_case_api.xlsx","login") # 变量接受函数的返回值


#发送接口请求
def api_request(api_url,data):

    qcd_headers = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    response = requests.post(url=api_url,json=data,headers=qcd_headers) # 返回值---响应消息
    return response.json()  #得到的响应结果

#数据回写
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename) # 加载工作簿对象
    sheet = wb[sheetname]   # 获取到表单
    sheet.cell(row=row,column=column).value = final_result # 数据重写
    wb.save(filename)  # 保存后才能写入生效

#数据对比
def excute_func(filename,sheetname):
    cases = read_data(filename,sheetname)
    for case in cases:
        case_id = case.get("case_id")
        url = case.get("url")
        data = eval(case["data"])
        expected = eval(case["expected"])
        real_result = api_request(api_url=url,data=data) # 执行结果
        real_msg = real_result["msg"]
        expected_msg = expected["msg"]
        print("执行结果是：{}".format(real_msg))
        print("预期结果是：{}".format(expected_msg))
        if expected_msg == real_msg:
            print("第{}条测试用例执行通过！".format(case_id))
            final_result = "Passed"
        else:
            print("第{}条测试用例执行不通过！".format(case_id))
            final_result = "Failed"
        print("*" * 20)
        write_result("test_case_api.xlsx","login",row=case_id+1,column=8,final_result=final_result)

excute_func("test_case_api.xlsx","login")
